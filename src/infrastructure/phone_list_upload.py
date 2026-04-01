"""Разбор CSV/XLSX со списком телефонов (первая колонка)."""

from __future__ import annotations

import csv
import io

from starlette.datastructures import UploadFile


async def extract_phones_from_upload(upload: UploadFile) -> list[str]:
    """Читает тело загрузки; для .xlsx — первая колонка листа, иначе CSV."""
    raw = await upload.read()
    if not raw:
        return []
    name = (upload.filename or "").lower()
    if name.endswith(".xlsx"):
        from openpyxl import load_workbook  # noqa: PLC0415 — тяжёлый импорт по необходимости

        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            ws = wb.active
            out: list[str] = []
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                cell = row[0] if row else None
                if cell is not None and str(cell).strip():
                    out.append(str(cell).strip())
            return out
        finally:
            wb.close()
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    return [row[0].strip() for row in reader if row and row[0].strip()]
