"""Разбор TXT и XLSX для загрузки в базу знаний (RAG)."""

from __future__ import annotations

import re
from io import BytesIO
from typing import Final

from openpyxl import load_workbook

# Ограничение на размер одного фрагмента для эмбеддинга (токены ~ на порядок меньше длины в байтах для RU).
_MAX_CHUNK_CHARS: Final[int] = 6000
_MAX_TITLE_LEN: Final[int] = 500


def decode_text_bytes(raw: bytes) -> str:
    """UTF-8, иначе типичная кириллица в Windows."""
    try:
        return raw.decode("utf-8")
    except UnicodeDecodeError:
        return raw.decode("cp1251", errors="replace")


def sanitize_base_title(filename: str) -> str:
    base = re.sub(r"[^\w\s\-.А-Яа-яЁё()]+", " ", filename, flags=re.UNICODE)
    base = re.sub(r"\s+", " ", base).strip()
    if not base:
        return "Документ"
    return base[:_MAX_TITLE_LEN]


def chunk_text(text: str, *, max_chars: int = _MAX_CHUNK_CHARS) -> list[str]:
    text = (text or "").strip()
    if not text:
        return []
    if len(text) <= max_chars:
        return [text]
    chunks: list[str] = []
    start = 0
    n = len(text)
    while start < n:
        end = min(start + max_chars, n)
        if end < n:
            nl = text.rfind("\n", start, end)
            if nl > start + max_chars // 2:
                end = nl + 1
        piece = text[start:end].strip()
        if piece:
            chunks.append(piece)
        start = end if end > start else start + max_chars
    return chunks


def ingest_txt(*, raw: bytes, filename: str) -> list[tuple[str, str]]:
    """Возвращает пары (заголовок, текст) для сохранения в БД."""
    body = decode_text_bytes(raw).strip()
    if not body:
        return []
    base = sanitize_base_title(filename.rsplit("/", maxsplit=1)[-1].rsplit("\\", maxsplit=1)[-1])
    if base.lower().endswith(".txt"):
        base = base[:-4].strip() or "Текст"
    chunks = chunk_text(body)
    out: list[tuple[str, str]] = []
    for i, chunk in enumerate(chunks):
        suffix = f" (часть {i + 1})" if len(chunks) > 1 else ""
        title = f"{base}{suffix}"[:_MAX_TITLE_LEN]
        out.append((title, chunk))
    return out


def _sheet_to_text(sheet) -> str:
    lines: list[str] = []
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if not any(cells):
            continue
        lines.append("\t".join(cells))
    return "\n".join(lines).strip()


def ingest_xlsx(*, raw: bytes, filename: str) -> list[tuple[str, str]]:
    """Каждый лист → один или несколько фрагментов с префиксом имени файла и листа."""
    base_file = sanitize_base_title(
        filename.rsplit("/", maxsplit=1)[-1].rsplit("\\", maxsplit=1)[-1]
    )
    if base_file.lower().endswith(".xlsx"):
        base_file = base_file[:-5].strip() or "Прайс"
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    out: list[tuple[str, str]] = []
    try:
        for sheet in wb.worksheets:
            text = _sheet_to_text(sheet)
            if not text:
                continue
            sheet_name = (sheet.title or "Лист").strip() or "Лист"
            prefix = f"{base_file} — {sheet_name}"
            chunks = chunk_text(text)
            for i, chunk in enumerate(chunks):
                suffix = f" (часть {i + 1})" if len(chunks) > 1 else ""
                title = f"{prefix}{suffix}"[:_MAX_TITLE_LEN]
                out.append((title, chunk))
    finally:
        wb.close()
    return out


def ingest_upload(*, raw: bytes, filename: str) -> list[tuple[str, str]]:
    name = filename.lower()
    if name.endswith(".txt"):
        return ingest_txt(raw=raw, filename=filename)
    if name.endswith(".xlsx"):
        return ingest_xlsx(raw=raw, filename=filename)
    msg = "Поддерживаются только .txt и .xlsx"
    raise ValueError(msg)
